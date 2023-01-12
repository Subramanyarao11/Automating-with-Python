import openpyxl
workbook = openpyxl.load_workbook('example.xlsx')
# print(type(workbook))
# <class 'openpyxl.workbook.workbook.Workbook'>
sheet1 = workbook["Sheet1"]
# print(type(sheet1))
sheets = workbook.sheetnames
# print(type(sheets))
cell = sheet1['B1']
# print(cell.value) // Prints 'Apples'


# Getting value of item in 1st row 2nd column
newValue = sheet1.cell(row=1, column=2).value
# print(newValue) // Prints 'Apples'
